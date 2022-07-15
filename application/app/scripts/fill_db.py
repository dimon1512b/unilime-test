import os
from pathlib import Path
from typing import List

import openpyxl
from sqlalchemy import insert, create_engine
from sqlalchemy.orm import Session

from app.config import settings
from app.db.models import Product, Review

ROOT_DIR = os.path.dirname(Path(__file__).parent.parent.parent)

products_file = os.path.join(ROOT_DIR, 'products.xlsx')
reviews_file = os.path.join(ROOT_DIR, 'reviews.xlsx')

wook_book_products = openpyxl.load_workbook(products_file)
wook_book_reviews = openpyxl.load_workbook(reviews_file)

sheet_products = wook_book_products.active
sheet_reviews = wook_book_reviews.active

user = settings.DB_USER
password = settings.DB_PASSWORD
db_name = settings.DB_NAME

engine = create_engine(
    f'postgresql://{user}:{password}@127.0.0.1:5432/{db_name}',
    echo=False,
    future=True,
)


def read_products():
    print('we in read_products')
    products_list = []
    for i in range(1, sheet_products.max_row):
        obj_to_append = {}
        for col in sheet_products.iter_cols(1, sheet_products.max_column):
            if col[i].column_letter == "A":
                obj_to_append["title"] = col[i].value
            elif col[i].column_letter == "B":
                obj_to_append["asin"] = col[i].value  # noqa
        products_list.append(obj_to_append)
    print(f'we out read_products\t{len(products_list) = }')
    return products_list


def read_reviews(products):
    print('we in read_reviews')
    products_asins = [product["asin"] for product in products]  # noqa
    reviews_list = []
    for i in range(1, sheet_reviews.max_row):
        obj_to_append = {}
        for col in sheet_reviews.iter_cols(1, sheet_reviews.max_column):
            if col[i].column_letter == "A":
                if col[i].value not in products_asins:
                    obj_to_append = {}
                    break
                obj_to_append["product_asin"] = col[i].value  # noqa
            elif col[i].column_letter == "B":
                obj_to_append["title"] = col[i].value
            elif col[i].column_letter == "C":
                obj_to_append["text"] = col[i].value
        reviews_list.append(obj_to_append) if obj_to_append else None
    print(f'we out read_reviews\t{len(reviews_list) = }')
    return reviews_list


def save_products(
    products: List[dict],
    session: Session,
):
    session.execute(
        insert(Product),
        products,
    )


def save_reviews(
    reviews: List[dict],
    session: Session,
):
    session.execute(
        insert(Review),
        reviews,
    )


def main():
    print('--- start')
    products = read_products()
    reviews = read_reviews(products)
    with Session(engine) as session:
        save_products(products, session)
        save_reviews(reviews, session)
        session.commit()
    print('--- finish')


if __name__ == '__main__':
    main()
